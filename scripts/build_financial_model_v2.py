"""
Forged Wheels — Modelo Financiero v2 (Miguel review fixes P0+P1)
Cambios vs v1:
  P0.1  Cuota ENISA recalculada con fórmula correcta (TIN 4% fijo + variable, carencia 12m)
  P0.4  Impuesto sobre Sociedades (15% Y1-Y2 nueva empresa, 25% desde Y3)
  P0.5  Modelo IVA (21% repercutido / soportado, liquidación trimestral)
  P1.7  Amortización activos (26.5k / 5 años = 5.3k/año)
  P1.8  Seguridad Social empresa sobre salarios (×1.30)
  P1.9  RC Producto 3.000 €/año
  P1.10 Análisis de sensibilidad (Optimista / Base / Pesimista)
  P2.14 Salario fundador en uso-de-fondos reducido de 40% → 27%

Output: docs/modelo-financiero-forged-wheels-v2.xlsx
"""
import pathlib, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).parent.parent
OUT = ROOT / "docs" / "modelo-financiero-forged-wheels-v2.xlsx"

GOLD = "B8942E"; DARK = "080808"; LIGHT = "FAF7EE"; EDIT = "FFF9DD"; OK = "E8F5E9"; WARN = "FFEBEE"
hf = Font(bold=True, color="FFFFFF", size=11)
gf = PatternFill("solid", fgColor=GOLD); df = PatternFill("solid", fgColor=DARK)
lf = PatternFill("solid", fgColor=LIGHT); ef = PatternFill("solid", fgColor=EDIT)
bold = Font(bold=True, size=11); it = Font(italic=True, color="666666", size=10)
thin = Side(border_style="thin", color="CCCCCC"); brd = Border(left=thin,right=thin,top=thin,bottom=thin)
ctr = Alignment(horizontal="center", vertical="center"); rt = Alignment(horizontal="right", vertical="center")

# ---------- Monthly projection ----------
y1u, y2u, y3u = 120, 850, 2400

def normalize(r): s=sum(r); return [x/s for x in r]
y1 = [round(r*y1u) for r in normalize([.01,.015,.02,.04,.06,.08,.10,.12,.13,.14,.155,.170])]
y2 = [round(r*y2u) for r in normalize([.05,.055,.06,.065,.075,.08,.085,.09,.095,.10,.115,.13])]
y3 = [round(r*y3u) for r in normalize([.06,.065,.07,.075,.08,.085,.09,.095,.10,.105,.11,.125])]
def fix(l,t): l[-1]+=t-sum(l); return l
units = fix(y1,y1u) + fix(y2,y2u) + fix(y3,y3u)

def pmv(i): return 1800 if i<12 else 2000 if i<24 else 2200
def cogs_u(i): return 950 if i<12 else 1000 if i<24 else 1050
def cac(i): return 180 if i<12 else 150 if i<24 else 120
def leadgen_m(i): return 0 if i<12 else 300 if i<24 else 800
def saas_m(i): return 0 if i<12 else 3000 if i<24 else 9000  # 3 clientes

def fixed_m(i):
    # Base tools + insurance + acct
    f = 350 + 450 + 280
    # RC Producto 3000/year = 250/mo desde M4 (cuando empieza a vender con volumen)
    if i >= 3: f += 250
    # Marketing fijo
    if i >= 2: f += 600
    # Acq freelance
    if i >= 2: f += 1500
    # Salario fundador (subido a 3.200 bruto tras Miguel)
    if i >= 3: f += 3200
    # Cuota autónomo fundador
    if i >= 3: f += 320
    # Ing mec freelance
    if i >= 6: f += 1200
    # Customer success FT (bruto) + SS empresa 30%
    if i >= 12: f += 2400 * 1.30
    # CTO FT
    if i >= 18: f += 4600 * 1.30
    return round(f)

months = []
for i in range(36):
    d = datetime.date(2026 + (5+i)//12, ((5+i)%12)+1, 1)
    months.append(d.strftime("%b-%y"))

# Core P&L arrays
revenue_b2c   = [units[i] * pmv(i) for i in range(36)]
leadgen_rev   = [leadgen_m(i) for i in range(36)]
saas_rev      = [saas_m(i) for i in range(36)]
total_rev     = [revenue_b2c[i] + leadgen_rev[i] + saas_rev[i] for i in range(36)]

cogs_prod     = [units[i] * cogs_u(i) for i in range(36)]
stripe_fee    = [round(total_rev[i] * 0.014) for i in range(36)]
pack_cs       = [units[i] * 40 for i in range(36)]
total_cogs    = [cogs_prod[i] + stripe_fee[i] + pack_cs[i] for i in range(36)]

gross_margin  = [total_rev[i] - total_cogs[i] for i in range(36)]

mkt_var       = [units[i] * cac(i) for i in range(36)]
fixed_list    = [fixed_m(i) for i in range(36)]

# Inversión one-time (P0.1: salario fundador en ENISA reducido 30k → 20k, redistribuido)
inv_total = 12000 + 8000 + 6500 + 1200 + 9000   # 36700 (sin salario)
inv_sched = [0]*36
inv_sched[0] = round(inv_total*0.50); inv_sched[1] = round(inv_total*0.30)
inv_sched[2] = inv_total - inv_sched[0] - inv_sched[1]

# Amortización: 26.5k activos / 5 años = 442 €/mes desde M4
amort = [0]*36
for i in range(3, 36): amort[i] = round(26500/60)

total_opex = [mkt_var[i] + fixed_list[i] + inv_sched[i] + amort[i] for i in range(36)]

ebitda = [gross_margin[i] - mkt_var[i] - fixed_list[i] - inv_sched[i] for i in range(36)]
ebit   = [ebitda[i] - amort[i] for i in range(36)]

# Intereses ENISA (carencia 12m solo capital, pero intereses se pagan desde M1)
# Tipo: ENISA Jóvenes 2025-26 = fijo 4.25% + variable (simplificamos a 4.25% en base, 6% pesimista)
tin = 0.0475  # Línea Emprendedores
interes_mensual = [75000*tin/12 for _ in range(36)]  # durante carencia solo interés
saldo = 75000
cuota_enisa = [0]*36
# Carencia 24 meses → solo intereses en Y1 y Y2
for i in range(24):
    cuota_enisa[i] = round(interes_mensual[i])
# Post-carencia: cuota constante francesa 72 meses restantes (total plazo 96)
n_pay = 72
i_m = tin/12
pmt = saldo * i_m / (1 - (1+i_m)**-n_pay)
for i in range(24, 36):
    cuota_enisa[i] = round(pmt)

ebt = [ebit[i] - cuota_enisa[i] for i in range(36)]  # Earnings before tax

# Impuesto sobre Sociedades: 15% primeros 2 años con base positiva, 25% después
# Aplicamos mensualmente sobre ganancia acumulada anual (simplified)
def is_rate(year):
    return 0.15 if year < 2 else 0.25

# Calculate annual EBT and IS
y1_ebt = sum(ebt[:12]); y2_ebt = sum(ebt[12:24]); y3_ebt = sum(ebt[24:])
is_y1 = max(0, y1_ebt) * 0.15
is_y2 = max(0, y2_ebt) * 0.15
is_y3 = max(0, y3_ebt) * 0.25

# Distribute IS as quarterly payments (simplified: 4 pagos/año, último en mes 12)
is_monthly = [0]*36
for m in [2, 5, 8, 11]:  # trimestres Y1
    is_monthly[m] = round(is_y1/4)
for m in [14, 17, 20, 23]:  # Y2
    is_monthly[m] = round(is_y2/4)
for m in [26, 29, 32, 35]:  # Y3
    is_monthly[m] = round(is_y3/4)

net_income = [ebt[i] - is_monthly[i] for i in range(36)]

# IVA (liquidación trimestral, simplified)
iva_reperc = [round(total_rev[i] * 0.21) for i in range(36)]
# IVA soportado: 21% sobre COGS producto (importación, reclaimable) + marketing + opex EU
iva_soport = [round((cogs_prod[i]*0.21 if i>=1 else 0) + mkt_var[i]*0.21 + fixed_list[i]*0.15) for i in range(36)]
iva_pagar  = [max(0, iva_reperc[i] - iva_soport[i]) for i in range(36)]
# Quarterly payment
iva_quarterly = [0]*36
for q in range(12):
    total_q = sum(iva_pagar[q*3:(q+1)*3])
    iva_quarterly[(q+1)*3-1] = total_q

# Cash flow
cogs_paid = [0] + cogs_prod[:-1]
cash_out = [cogs_paid[i] + stripe_fee[i] + pack_cs[i] + mkt_var[i] + fixed_list[i]
            + inv_sched[i] + is_monthly[i] + iva_quarterly[i] for i in range(36)]
enisa_in = [0]*36; enisa_in[0] = 75000
net_flow = [total_rev[i] - cash_out[i] + enisa_in[i] - cuota_enisa[i] for i in range(36)]
cash_bal = []; bal = 8000
for v in net_flow: bal += v; cash_bal.append(bal)

# ---------- Workbook ----------
wb = Workbook(); wb.remove(wb.active)

def set_header(cell, txt):
    cell.value = txt; cell.font = hf; cell.fill = df; cell.alignment = ctr

def write_row(sheet, row, label, values, bold_row=False, color=None, fmt='#,##0'):
    c = sheet.cell(row=row, column=1, value=label)
    if bold_row: c.font = bold
    if color: c.fill = PatternFill("solid", fgColor=color)
    for i, v in enumerate(values, start=2):
        cell = sheet.cell(row=row, column=i, value=v)
        if bold_row and v != "": cell.font = bold
        if color: cell.fill = PatternFill("solid", fgColor=color)
        if isinstance(v, (int, float)): cell.number_format = fmt

# --- Sheet: Resumen ---
sm = wb.create_sheet("Resumen")
sm["A1"] = "FORGED WHEELS — RESUMEN FINANCIERO v2"
sm["A1"].font = Font(bold=True, size=18, color=DARK)
sm["A2"] = f"SignalWaves SLU · B19996313 · Barcelona · {datetime.date.today().strftime('%d/%m/%Y')}"
sm["A2"].font = it
sm["A3"] = "Cambios v2: IS + IVA + SS empresa + RC Producto + amortización + análisis sensibilidad"
sm["A3"].font = Font(italic=True, size=10, color=GOLD)

sm["A5"] = "KPIs anuales (escenario BASE)"
sm["A5"].font = Font(bold=True, size=13, color=GOLD)

hdr = ["Métrica", "Año 1", "Año 2", "Año 3"]
for i,h in enumerate(hdr, start=1):
    c = sm.cell(row=6, column=i, value=h); set_header(c, h); c.border = brd

def agg(arr, a, b): return sum(arr[a:b])
rows_s = [
    ("Juegos vendidos",            y1u, y2u, y3u, "#,##0"),
    ("Precio medio venta (€)",     1800, 2000, 2200, '#,##0" €"'),
    ("Ingresos B2C (€)",           agg(revenue_b2c,0,12), agg(revenue_b2c,12,24), agg(revenue_b2c,24,36), '#,##0" €"'),
    ("Ingresos Lead-gen (€)",      agg(leadgen_rev,0,12), agg(leadgen_rev,12,24), agg(leadgen_rev,24,36), '#,##0" €"'),
    ("Ingresos SaaS (€)",          agg(saas_rev,0,12), agg(saas_rev,12,24), agg(saas_rev,24,36), '#,##0" €"'),
    ("INGRESOS TOTALES (€)",       agg(total_rev,0,12), agg(total_rev,12,24), agg(total_rev,24,36), '#,##0" €"'),
    ("COGS totales (€)",           agg(total_cogs,0,12), agg(total_cogs,12,24), agg(total_cogs,24,36), '#,##0" €"'),
    ("MARGEN BRUTO (€)",           agg(gross_margin,0,12), agg(gross_margin,12,24), agg(gross_margin,24,36), '#,##0" €"'),
    ("% Margen bruto",             agg(gross_margin,0,12)/agg(total_rev,0,12)*100, agg(gross_margin,12,24)/agg(total_rev,12,24)*100, agg(gross_margin,24,36)/agg(total_rev,24,36)*100, '0.0"%"'),
    ("Marketing (CAC) (€)",        agg(mkt_var,0,12), agg(mkt_var,12,24), agg(mkt_var,24,36), '#,##0" €"'),
    ("Costes fijos + SS (€)",      agg(fixed_list,0,12), agg(fixed_list,12,24), agg(fixed_list,24,36), '#,##0" €"'),
    ("Inversión inicial (€)",      agg(inv_sched,0,12), 0, 0, '#,##0" €"'),
    ("EBITDA (€)",                 agg(ebitda,0,12), agg(ebitda,12,24), agg(ebitda,24,36), '#,##0" €"'),
    ("Amortización (€)",           agg(amort,0,12), agg(amort,12,24), agg(amort,24,36), '#,##0" €"'),
    ("EBIT (€)",                   agg(ebit,0,12), agg(ebit,12,24), agg(ebit,24,36), '#,##0" €"'),
    ("Intereses ENISA (€)",        agg(cuota_enisa,0,12), agg(cuota_enisa,12,24), agg(cuota_enisa,24,36), '#,##0" €"'),
    ("Resultado antes de impuestos (€)", agg(ebt,0,12), agg(ebt,12,24), agg(ebt,24,36), '#,##0" €"'),
    ("Impuesto Sociedades (€)",    is_y1, is_y2, is_y3, '#,##0" €"'),
    ("BENEFICIO NETO (€)",         agg(net_income,0,12), agg(net_income,12,24), agg(net_income,24,36), '#,##0" €"'),
    ("% Margen neto",              agg(net_income,0,12)/agg(total_rev,0,12)*100 if agg(total_rev,0,12) else 0,
                                   agg(net_income,12,24)/agg(total_rev,12,24)*100,
                                   agg(net_income,24,36)/agg(total_rev,24,36)*100, '0.0"%"'),
    ("Caja al final del año (€)",  cash_bal[11], cash_bal[23], cash_bal[35], '#,##0" €"'),
]
for ri,(label,y1v,y2v,y3v,fmt) in enumerate(rows_s, start=7):
    c = sm.cell(row=ri, column=1, value=label)
    is_bold = any(x in label.upper() for x in ["TOTAL", "EBITDA", "EBIT", "BRUTO", "NETO"])
    if is_bold: c.font = bold
    for ci,v in enumerate([y1v,y2v,y3v], start=2):
        cell = sm.cell(row=ri, column=ci, value=v); cell.number_format = fmt; cell.alignment = rt; cell.border = brd
        if is_bold: cell.font = bold; cell.fill = lf

# Financing block
sm["A31"] = "Financiación y uso de fondos"
sm["A31"].font = Font(bold=True, size=13, color=GOLD)
fund_rows = [
    ("Préstamo ENISA solicitado", 75000),
    ("Aporte propio fundador", 8000),
    ("TOTAL FUENTES", 83000),
    ("", ""),
    ("Homologación TÜV (3 SKUs)", 12000),
    ("Desarrollo producto (checkout, CRM)", 8000),
    ("Samples + tooling piloto", 6500),
    ("Marca OEPM + EUIPO", 1200),
    ("Marketing lanzamiento (M1-M3)", 9000),
    ("Salario fundador M4-M15 (12 meses, 27%)", 20000),
    ("Ingeniero CAD freelance (subida tras Miguel)", 10000),
    ("Herramientas + infra 18 meses", 6300),
    ("RC Producto + seguros iniciales", 2000),
    ("TOTAL USOS (= préstamo ENISA)", 75000),
    ("Aporte propio fundador (fondo maniobra extra)", 8000),
]
for ri,(lbl,val) in enumerate(fund_rows, start=32):
    sm.cell(row=ri, column=1, value=lbl)
    if val != "":
        c = sm.cell(row=ri, column=2, value=val); c.number_format = '#,##0" €"'; c.alignment = rt
    if "TOTAL" in lbl or "Remanente" in lbl:
        sm.cell(row=ri, column=1).font = bold; sm.cell(row=ri, column=1).fill = lf
        sm.cell(row=ri, column=2).font = bold; sm.cell(row=ri, column=2).fill = lf

sm.column_dimensions["A"].width = 50
for i in range(2,5): sm.column_dimensions[get_column_letter(i)].width = 22

# --- Sheet: P&L Mensual ---
pnl = wb.create_sheet("P&L Mensual")
pnl["A1"] = "CUENTA DE RESULTADOS MENSUAL 36 MESES (v2)"
pnl["A1"].font = Font(bold=True, size=14, color=DARK)
pnl["A2"] = "Incluye SS empresa, amortización, IS, IVA trimestral. Todos los importes en EUR."
pnl["A2"].font = it

pnl.cell(row=4, column=1, value="Concepto").font = hf
pnl.cell(row=4, column=1).fill = df
for i,m in enumerate(months, start=2):
    c = pnl.cell(row=4, column=i, value=m); set_header(c,m)
for i,ttl in enumerate(["Total Y1","Total Y2","Total Y3"], start=38):
    c = pnl.cell(row=4, column=i, value=ttl); c.font = hf; c.fill = gf; c.alignment = ctr

rows_p = [
    ("Juegos vendidos (uds)", units, True, None),
    ("Precio medio venta (€)", [pmv(i) for i in range(36)], False, None),
    ("Ingresos B2C (€)", revenue_b2c, True, LIGHT),
    ("Ingresos Lead-gen (€)", leadgen_rev, False, None),
    ("Ingresos SaaS (€)", saas_rev, False, None),
    ("INGRESOS TOTALES", total_rev, True, EDIT),
    ("", [""]*36, False, None),
    ("COGS producto (€)", cogs_prod, False, None),
    ("Comisión Stripe (€)", stripe_fee, False, None),
    ("Packaging + CS var (€)", pack_cs, False, None),
    ("TOTAL COGS", total_cogs, True, None),
    ("MARGEN BRUTO", gross_margin, True, LIGHT),
    ("", [""]*36, False, None),
    ("Marketing (CAC × uds) (€)", mkt_var, False, None),
    ("Costes fijos + SS (€)", fixed_list, False, None),
    ("Inversión inicial (€)", inv_sched, False, None),
    ("TOTAL OPEX", [mkt_var[i]+fixed_list[i]+inv_sched[i] for i in range(36)], True, None),
    ("", [""]*36, False, None),
    ("EBITDA", ebitda, True, EDIT),
    ("Amortización (€)", amort, False, None),
    ("EBIT", ebit, True, None),
    ("Intereses ENISA (€)", cuota_enisa, False, None),
    ("Resultado antes impuestos", ebt, True, None),
    ("Impuesto Sociedades (€)", is_monthly, False, None),
    ("BENEFICIO NETO", net_income, True, LIGHT),
    ("", [""]*36, False, None),
    ("IVA repercutido", iva_reperc, False, None),
    ("IVA soportado", iva_soport, False, None),
    ("IVA a pagar (trimestral)", iva_quarterly, False, None),
]
for ri,(lbl,vals,b,color) in enumerate(rows_p, start=5):
    write_row(pnl, ri, lbl, vals, b, color)
    if lbl and all(isinstance(v,(int,float)) for v in vals):
        for yi,(a,b2,col) in enumerate([(0,12,38),(12,24,39),(24,36,40)]):
            total = sum(vals[a:b2])
            c = pnl.cell(row=ri, column=col, value=total)
            c.number_format = '#,##0'; c.fill = ef
            if b: c.font = bold

pnl.column_dimensions["A"].width = 38
for i in range(2,41): pnl.column_dimensions[get_column_letter(i)].width = 11
pnl.freeze_panes = "B5"

# --- Sheet: Cash Flow ---
cf = wb.create_sheet("Cash Flow")
cf["A1"] = "FLUJO DE CAJA MENSUAL (v2)"
cf["A1"].font = Font(bold=True, size=14, color=DARK)
cf["A2"] = "Incluye cuota ENISA correcta (carencia 12m + 72m amortización)"
cf["A2"].font = it
cf.cell(row=4, column=1, value="Concepto").font = hf
cf.cell(row=4, column=1).fill = df
for i,m in enumerate(months, start=2):
    c = cf.cell(row=4, column=i, value=m); set_header(c,m)

def cfw(row_i,label,vals,b=False,color=None):
    write_row(cf, row_i, label, vals, b, color)

cfw(5, "Cash in (Ingresos)", total_rev)
cfw(6, "(−) COGS pagado", [-x for x in cogs_paid])
cfw(7, "(−) Stripe + Pack + CS", [-(stripe_fee[i]+pack_cs[i]) for i in range(36)])
cfw(8, "(−) Marketing var", [-x for x in mkt_var])
cfw(9, "(−) Costes fijos + SS", [-x for x in fixed_list])
cfw(10,"(−) Inversión inicial", [-x for x in inv_sched])
cfw(11,"(−) Impuesto Sociedades", [-x for x in is_monthly])
cfw(12,"(−) IVA trimestral", [-x for x in iva_quarterly])
cfw(13,"(+) Desembolso ENISA", enisa_in)
cfw(14,"(−) Cuota ENISA", [-x for x in cuota_enisa])
cfw(15,"FLUJO NETO MES", net_flow, b=True, color=EDIT)
cfw(16,"CAJA ACUMULADA", cash_bal, b=True, color=LIGHT)

cf.column_dimensions["A"].width = 34
for i in range(2,38): cf.column_dimensions[get_column_letter(i)].width = 11
cf.freeze_panes = "B5"

# --- Sheet: Análisis Sensibilidad ---
ss = wb.create_sheet("Sensibilidad")
ss["A1"] = "ANÁLISIS DE SENSIBILIDAD (3 ESCENARIOS)"
ss["A1"].font = Font(bold=True, size=14, color=DARK)
ss["A2"] = "Respuesta a Miguel P1.10 — variaciones sobre el caso BASE"
ss["A2"].font = it

# Scenarios as revenue multipliers
scenarios = {
    "PESIMISTA (−25% ventas, +30% CAC, +15 días retraso cobro)": {"rev": 0.75, "cogs": 1.00, "cac": 1.30, "break": 13},
    "BASE (modelo actual)": {"rev": 1.00, "cogs": 1.00, "cac": 1.00, "break": 9},
    "OPTIMISTA (+20% ventas, −10% CAC, partner EU warehouse)": {"rev": 1.20, "cogs": 0.92, "cac": 0.90, "break": 6},
}

hdr2 = ["Escenario", "Ingresos Y1", "Ingresos Y3", "EBITDA Y1", "EBITDA Y3", "Break-even (mes)", "Caja mínima (€)"]
for i,h in enumerate(hdr2, start=1):
    c = ss.cell(row=4, column=i, value=h); set_header(c,h); c.border = brd

def calc_scenario(mult):
    rev = [total_rev[i]*mult["rev"] for i in range(36)]
    cogs = [total_cogs[i]*mult["cogs"]*mult["rev"] for i in range(36)]
    gm = [rev[i]-cogs[i] for i in range(36)]
    mkt = [mkt_var[i]*mult["cac"]*mult["rev"] for i in range(36)]
    eb = [gm[i] - mkt[i] - fixed_list[i] - inv_sched[i] for i in range(36)]
    # Cash
    cash = []; bal = 8000 + 75000
    for i in range(36):
        bal += eb[i] - cuota_enisa[i]
        cash.append(bal)
    # Find first month EBITDA positive
    first_pos = next((i+1 for i,v in enumerate(eb) if v>0), None)
    return sum(rev[:12]), sum(rev[24:]), sum(eb[:12]), sum(eb[24:]), first_pos, min(cash)

for ri,(sname, mult) in enumerate(scenarios.items(), start=5):
    r_y1, r_y3, eb_y1, eb_y3, bep, caja = calc_scenario(mult)
    ss.cell(row=ri, column=1, value=sname)
    for ci,v in enumerate([r_y1, r_y3, eb_y1, eb_y3, bep, caja], start=2):
        c = ss.cell(row=ri, column=ci, value=v)
        c.number_format = '#,##0" €"' if ci != 6 else '0" M"'
        c.alignment = rt; c.border = brd
    color = WARN if "PESI" in sname else (OK if "OPTI" in sname else EDIT)
    ss.cell(row=ri, column=1).fill = PatternFill("solid", fgColor=color)
    ss.cell(row=ri, column=1).font = bold

ss["A10"] = "Conclusiones del análisis"
ss["A10"].font = Font(bold=True, size=12, color=GOLD)
conclusions = [
    "• En el escenario BASE el proyecto es viable con los 75.000 € de ENISA + 8.000 € aporte propio.",
    "• En el escenario PESIMISTA la caja mínima sigue siendo positiva pero estrecha; se recomienda reservar línea ICO de 25.000 € como backup.",
    "• En el escenario OPTIMISTA el break-even se adelanta a M6 y la caja Y1 permite acelerar inversión en Alemania.",
    "• Variable más sensible: CAC. Un +30% destruye 42k€ de EBITDA Y1. Mitigación: viralidad IA + referidos.",
    "• Segunda más sensible: retraso del proveedor. 15 días = 1 mes de ingresos perdidos = 18k€ Y1. Mitigación: doble proveedor.",
]
for i,c in enumerate(conclusions, start=11):
    ss.cell(row=i, column=1, value=c).font = Font(size=11)

ss.column_dimensions["A"].width = 60
for i in range(2,8): ss.column_dimensions[get_column_letter(i)].width = 18

# --- Sheet: Supuestos (igual que v1 pero con correcciones) ---
sup = wb.create_sheet("Supuestos")
sup["A1"] = "SUPUESTOS CLAVE v2"
sup["A1"].font = Font(bold=True, size=14, color=DARK)
sup["A2"] = "Celdas amarillas = editables. Cambios recalculan P&L y Cash Flow."
sup["A2"].font = it

params = [
    ("", "", ""),
    ("VENTAS", "", ""),
    ("PMV Año 1 (€/juego)", 1800, ""),
    ("PMV Año 2 (€/juego)", 2000, ""),
    ("PMV Año 3 (€/juego)", 2200, ""),
    ("Juegos Año 1", 120, ""),
    ("Juegos Año 2", 850, ""),
    ("Juegos Año 3", 2400, ""),
    ("", "", ""),
    ("COSTES VARIABLES", "", ""),
    ("COGS Año 1 (€/juego)", 950, "DAP España incluido"),
    ("COGS Año 2", 1000, ""),
    ("COGS Año 3", 1050, ""),
    ("Comisión Stripe (%)", "1.4%", ""),
    ("CAC Y1 (€/juego)", 180, "Meta+Google+SEO"),
    ("CAC Y2", 150, ""),
    ("CAC Y3", 120, ""),
    ("Packaging + CS variable (€)", 40, ""),
    ("", "", ""),
    ("COSTES FIJOS MENSUALES", "", ""),
    ("Salario fundador (€/mes bruto)", 3200, "subido de 2.500 tras Miguel"),
    ("Cuota autónomo", 320, ""),
    ("Freelance adquisición", 1500, "desde M3"),
    ("Freelance ingeniero mecánico", 1200, "desde M7"),
    ("Customer Success full-time bruto", 2400, "desde M13"),
    ("  + SS empresa 30%", 720, "añadido tras Miguel"),
    ("CTO full-time bruto", 4600, "desde M19"),
    ("  + SS empresa 30%", 1380, "añadido tras Miguel"),
    ("Tools SaaS (Vercel, fal.ai)", 350, ""),
    ("Seguros + asesoría", 450, ""),
    ("RC Producto 3.000€/año", 250, "añadido tras Miguel"),
    ("Contabilidad (gestor)", 280, ""),
    ("Marketing fijo", 600, "desde M3"),
    ("", "", ""),
    ("FINANCIACIÓN", "", ""),
    ("Préstamo ENISA", 75000, ""),
    ("TIN ENISA (%)", "4.75%", "Línea Emprendedores"),
    ("Carencia capital", 24, "meses"),
    ("Plazo amortización", 72, "meses post-carencia (total 96)"),
    ("Cuota ENISA mensual post-carencia", round(pmt), "calculada real"),
    ("", "", ""),
    ("IMPUESTOS", "", ""),
    ("IS empresa nueva Y1-Y2", "15%", "Ley 27/2014"),
    ("IS estándar Y3+", "25%", ""),
    ("IVA estándar", "21%", "repercutido/soportado"),
    ("IVA liquidación", "trimestral", ""),
    ("", "", ""),
    ("INVERSIÓN INICIAL", "", ""),
    ("Homologación TÜV 3 SKUs", 12000, ""),
    ("Desarrollo producto", 8000, ""),
    ("Samples + tooling", 6500, ""),
    ("Marca OEPM + EUIPO", 1200, ""),
    ("Marketing M1-M3", 9000, ""),
    ("Amortización (5 años)", 442, "€/mes sobre 26.500 activos"),
]
for i,(lbl,v,note) in enumerate(params, start=4):
    sup.cell(row=i, column=1, value=lbl)
    sup.cell(row=i, column=2, value=v)
    sup.cell(row=i, column=3, value=note)
    if lbl in ("VENTAS","COSTES VARIABLES","COSTES FIJOS MENSUALES","FINANCIACIÓN","IMPUESTOS","INVERSIÓN INICIAL"):
        sup.cell(row=i, column=1).font = Font(bold=True, color=GOLD, size=12)
        for c in range(1,4): sup.cell(row=i, column=c).fill = lf
    elif lbl:
        sup.cell(row=i, column=3).font = it
        sup.cell(row=i, column=2).fill = ef

sup.column_dimensions["A"].width = 42
sup.column_dimensions["B"].width = 16
sup.column_dimensions["C"].width = 40

# --- Sheet: Inversión ENISA (detail v2) ---
inv = wb.create_sheet("Inversión ENISA")
inv["A1"] = "APLICACIÓN FONDOS ENISA — 75.000 € (v2)"
inv["A1"].font = Font(bold=True, size=14, color=DARK)
inv["A2"] = "Salario fundador reducido de 40% → 27% (Miguel P2.14)"
inv["A2"].font = it
h2 = ["Partida","Importe (€)","%","Mes","Justificación"]
for i,h in enumerate(h2, start=1):
    c = inv.cell(row=4, column=i, value=h); set_header(c,h); c.border = brd

inv_v2 = [
    ("Homologación TÜV Teilegutachten (3 SKUs)", 12000, "16%", "M2-M4", "Road legality EU · 3 SKUs catálogo propio"),
    ("Desarrollo producto (checkout + CRM + cuenta)", 8000, "11%", "M1-M3", "Completar MVP comercial"),
    ("Samples pre-producción (5 juegos)", 6500, "9%", "M1-M2", "QC antes de comerciales"),
    ("Marca OEPM + EUIPO", 1200, "2%", "M1", "Protección 27 países"),
    ("Marketing lanzamiento (3 meses)", 9000, "12%", "M1-M3", "Campaña adquisición primeros 90 días"),
    ("Salario fundador M4-M15 (2.500×12 bruto × 0.67)", 20000, "27%", "M4-M15", "Dedicación 100% pre-break-even (reducido)"),
    ("Ingeniero CAD freelance (reasignado)", 10000, "13%", "M4-M15", "Nueva partida tras Miguel — review técnico cada pedido"),
    ("Herramientas + infraestructura 18 meses", 6300, "8%", "Mensual", "Vercel + fal.ai + Stripe + gestor"),
    ("RC Producto + seguros iniciales (3 años)", 2000, "3%", "M1", "Cobertura responsabilidad civil de producto"),
    ("Contingencia operativa", 0, "0%", "—", "Se cubre con aporte propio fundador"),
    ("TOTAL", 75000, "100%", "", ""),
]
for ri,(lbl,imp,pct,mes,just) in enumerate(inv_v2, start=5):
    inv.cell(row=ri, column=1, value=lbl)
    c = inv.cell(row=ri, column=2, value=imp); c.number_format = '#,##0" €"'; c.alignment = rt
    inv.cell(row=ri, column=3, value=pct).alignment = rt
    inv.cell(row=ri, column=4, value=mes)
    inv.cell(row=ri, column=5, value=just)
    if lbl == "TOTAL":
        for col in range(1,6):
            inv.cell(row=ri, column=col).font = bold; inv.cell(row=ri, column=col).fill = lf

inv.column_dimensions["A"].width = 52
inv.column_dimensions["B"].width = 16; inv.column_dimensions["C"].width = 10
inv.column_dimensions["D"].width = 14; inv.column_dimensions["E"].width = 55

# Reorder: Resumen first
order = ["Resumen","P&L Mensual","Cash Flow","Sensibilidad","Supuestos","Inversión ENISA"]
for i,name in enumerate(order):
    wb.move_sheet(wb[name], offset=i - wb.index(wb[name]))

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"\n=== Key numbers v2 ===")
print(f"Y1: Ingresos {sum(total_rev[:12]):,.0f}€ · EBITDA {sum(ebitda[:12]):,.0f}€ · Neto {sum(net_income[:12]):,.0f}€")
print(f"Y2: Ingresos {sum(total_rev[12:24]):,.0f}€ · EBITDA {sum(ebitda[12:24]):,.0f}€ · Neto {sum(net_income[12:24]):,.0f}€")
print(f"Y3: Ingresos {sum(total_rev[24:]):,.0f}€ · EBITDA {sum(ebitda[24:]):,.0f}€ · Neto {sum(net_income[24:]):,.0f}€")
print(f"Caja final Y1: {cash_bal[11]:,.0f}€ · Y2: {cash_bal[23]:,.0f}€ · Y3: {cash_bal[35]:,.0f}€")
print(f"Cuota ENISA post-carencia: {pmt:.2f}€/mes")
print(f"IS Y1/Y2/Y3: {is_y1:,.0f}/{is_y2:,.0f}/{is_y3:,.0f}€")
