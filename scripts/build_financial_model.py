"""
Build the Forged Wheels financial model for ENISA.
Output: docs/modelo-financiero-forged-wheels-v1.xlsx
Sheets:
  - Resumen        — annual KPIs
  - Supuestos      — assumptions (editable by user, drives everything)
  - P&L Mensual    — 36-month profit & loss
  - Cash Flow      — 36-month cash
  - Inversión ENISA — use of funds
"""
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).parent.parent
OUT = ROOT / "docs" / "modelo-financiero-forged-wheels-v1.xlsx"

# Styles
GOLD = "B8942E"
DARK = "080808"
LIGHT_BG = "FAF7EE"
header_font = Font(bold=True, color="FFFFFF", size=11)
gold_fill = PatternFill("solid", fgColor=GOLD)
dark_fill = PatternFill("solid", fgColor=DARK)
light_fill = PatternFill("solid", fgColor=LIGHT_BG)
bold = Font(bold=True, size=11)
italic = Font(italic=True, color="666666", size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
right = Alignment(horizontal="right", vertical="center")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

# =====================================================================
# Sheet: Supuestos
# =====================================================================
s = wb.create_sheet("Supuestos")
s["A1"] = "SUPUESTOS CLAVE DEL MODELO FINANCIERO"
s["A1"].font = Font(bold=True, size=14, color=DARK)
s["A2"] = "Editar celdas amarillas para recalcular P&L y Cash Flow"
s["A2"].font = italic

headers = [
    ("", "", ""),
    ("VENTAS — B2C", "", ""),
    ("Precio medio venta Año 1 (PMV, €)", 1800, "por juego de 4 llantas"),
    ("Precio medio venta Año 2 (€)", 2000, "sube con catálogo premium"),
    ("Precio medio venta Año 3 (€)", 2200, ""),
    ("Juegos vendidos Año 1", 120, "M1-M12 desde lanzamiento"),
    ("Juegos vendidos Año 2", 850, ""),
    ("Juegos vendidos Año 3", 2400, ""),
    ("", "", ""),
    ("COSTES VARIABLES", "", ""),
    ("COGS medio Año 1 (€ por juego)", 950, "incluye llanta + flete FOB"),
    ("COGS medio Año 2 (€)", 1000, ""),
    ("COGS medio Año 3 (€)", 1050, ""),
    ("Comisión Stripe (%)", 0.014, "1.4% sobre venta"),
    ("CAC blended Año 1 (€)", 180, "Meta + Google + otros"),
    ("CAC blended Año 2 (€)", 150, "optimización"),
    ("CAC blended Año 3 (€)", 120, ""),
    ("Packaging + customer service (€/juego)", 40, ""),
    ("", "", ""),
    ("INGRESOS ADICIONALES", "", ""),
    ("Lead-gen: leads/mes Año 1", 0, "hasta 15% ingresos año 3"),
    ("Lead-gen: comisión/lead (€)", 200, ""),
    ("SaaS: clientes Año 2", 1, "fee mensual fijo"),
    ("SaaS: fee mensual/cliente (€)", 3000, ""),
    ("SaaS: clientes Año 3", 3, ""),
    ("", "", ""),
    ("COSTES FIJOS MENSUALES", "", ""),
    ("Salario fundador (€/mes bruto)", 2500, "desde M4"),
    ("Adquisición digital freelance (€/mes)", 1500, "desde M3"),
    ("Ingeniero mecánico freelance (€/mes)", 1200, "desde M7"),
    ("Customer success full-time (€/mes)", 2400, "desde M13"),
    ("CTO full-time (€/mes)", 4600, "desde M19"),
    ("Herramientas SaaS (Vercel, fal.ai, Stripe, etc.) (€/mes)", 350, ""),
    ("Oficina/coworking (€/mes)", 0, "remoto"),
    ("Seguros + asesoría jurídica (€/mes)", 450, ""),
    ("Contabilidad + fiscal (€/mes)", 280, "gestor"),
    ("Marketing fijo (contenido, SEO) (€/mes)", 600, "desde M3"),
    ("", "", ""),
    ("INVERSIÓN INICIAL (one-time)", "", ""),
    ("Homologación TÜV catálogo propio (€)", 12000, "3 SKUs iniciales"),
    ("Desarrollo producto (checkout, CRM) (€)", 8000, ""),
    ("Samples + tooling piloto (€)", 6500, "5 juegos sample"),
    ("Marca OEPM + EUIPO (€)", 1200, ""),
    ("Marketing lanzamiento (M1-M3 extra) (€)", 9000, ""),
    ("", "", ""),
    ("FINANCIACIÓN", "", ""),
    ("Préstamo ENISA (€)", 75000, "Línea Jóvenes Emprendedores"),
    ("TIN ENISA (%)", 0.06, "aprox 6%"),
    ("Carencia capital (meses)", 12, ""),
    ("Plazo total (meses)", 84, "7 años"),
    ("Aporte propio fundador (€)", 8000, ""),
]

for i, (label, val, note) in enumerate(headers, start=4):
    s.cell(row=i, column=1, value=label)
    s.cell(row=i, column=2, value=val)
    s.cell(row=i, column=3, value=note)
    if label in ("VENTAS — B2C", "COSTES VARIABLES", "INGRESOS ADICIONALES",
                 "COSTES FIJOS MENSUALES", "INVERSIÓN INICIAL (one-time)", "FINANCIACIÓN"):
        s.cell(row=i, column=1).font = Font(bold=True, color=GOLD, size=12)
        for col in range(1, 4):
            s.cell(row=i, column=col).fill = light_fill
    elif label:
        s.cell(row=i, column=1).font = Font(size=11)
        s.cell(row=i, column=3).font = italic
        s.cell(row=i, column=2).fill = PatternFill("solid", fgColor="FFF9DD")  # editable yellow

s.column_dimensions["A"].width = 55
s.column_dimensions["B"].width = 14
s.column_dimensions["C"].width = 35

# =====================================================================
# Sheet: P&L Mensual
# =====================================================================
pnl = wb.create_sheet("P&L Mensual")
pnl["A1"] = "CUENTA DE RESULTADOS — MENSUAL 36 MESES"
pnl["A1"].font = Font(bold=True, size=14, color=DARK)
pnl["A2"] = "Todos los importes en EUR. Año 1 = Jun 2026 – May 2027."
pnl["A2"].font = italic

# Unit plan: month-by-month ramp-up
y1_units = 120  # year 1
y2_units = 850
y3_units = 2400

# ramp curve (share of year)
y1_ramp = [0.01,0.015,0.02,0.04,0.06,0.08,0.10,0.12,0.13,0.14,0.155,0.170]  # sums ~1
y2_ramp = [0.05,0.055,0.06,0.065,0.075,0.08,0.085,0.09,0.095,0.10,0.115,0.13]
y3_ramp = [0.06,0.065,0.07,0.075,0.08,0.085,0.09,0.095,0.10,0.105,0.11,0.125]

def normalize(ramp):
    s_ = sum(ramp); return [r/s_ for r in ramp]

y1 = [round(r*y1_units) for r in normalize(y1_ramp)]
y2 = [round(r*y2_units) for r in normalize(y2_ramp)]
y3 = [round(r*y3_units) for r in normalize(y3_ramp)]

# Fix totals
def fix(l, t):
    diff = t - sum(l); l[-1] += diff; return l
y1 = fix(y1, y1_units)
y2 = fix(y2, y2_units)
y3 = fix(y3, y3_units)
all_units = y1 + y2 + y3

# Month labels (Jun 2026 - May 2029)
import datetime
start = datetime.date(2026, 6, 1)
months = []
for i in range(36):
    d = datetime.date(2026 + (5+i)//12, ((5+i)%12)+1, 1)
    months.append(d.strftime("%b-%y"))

# header row
pnl.cell(row=4, column=1, value="Concepto").font = header_font
pnl.cell(row=4, column=1).fill = dark_fill
for i, m in enumerate(months, start=2):
    c = pnl.cell(row=4, column=i, value=m)
    c.font = header_font; c.fill = dark_fill; c.alignment = center
pnl.cell(row=4, column=38, value="Total Y1").font = header_font
pnl.cell(row=4, column=38).fill = gold_fill
pnl.cell(row=4, column=39, value="Total Y2").font = header_font
pnl.cell(row=4, column=39).fill = gold_fill
pnl.cell(row=4, column=40, value="Total Y3").font = header_font
pnl.cell(row=4, column=40).fill = gold_fill

# Per-month values
def pmv(i):
    if i < 12: return 1800
    elif i < 24: return 2000
    else: return 2200
def cogs_unit(i):
    if i < 12: return 950
    elif i < 24: return 1000
    else: return 1050
def cac(i):
    if i < 12: return 180
    elif i < 24: return 150
    else: return 120
def leadgen_mo(i):
    if i < 12: return 0
    elif i < 24: return 300   # €/mes Y2
    else: return 800
def saas_mo(i):
    if i < 12: return 0
    elif i < 24: return 3000 * 1
    else: return 3000 * 3
def fixed(i):
    # ramp fixed costs
    f = 0
    if i >= 0: f += 350 + 450 + 280  # tools + seguros + contab
    if i >= 2: f += 600              # marketing fijo
    if i >= 2: f += 1500             # acq freelance
    if i >= 3: f += 2500             # fundador
    if i >= 6: f += 1200             # ing mec freelance
    if i >= 12: f += 2400            # CS FT
    if i >= 18: f += 4600            # CTO FT
    return f

rows_data = []

def row(label, values, bold_row=False, color=None):
    rows_data.append((label, values, bold_row, color))

# Ingresos
revenue = [all_units[i] * pmv(i) for i in range(36)]
row("Juegos vendidos (uds)", all_units, bold_row=True)
row("Precio medio venta (€)", [pmv(i) for i in range(36)])
row("Ingresos B2C (€)", revenue, bold_row=True, color=LIGHT_BG)
row("Ingresos Lead-gen (€)", [leadgen_mo(i) for i in range(36)])
row("Ingresos SaaS (€)", [saas_mo(i) for i in range(36)])
total_rev = [revenue[i] + leadgen_mo(i) + saas_mo(i) for i in range(36)]
row("INGRESOS TOTALES", total_rev, bold_row=True, color="FFF9DD")

# Espaciador
row("", ["" for _ in range(36)])

# COGS
cogs_prod = [all_units[i] * cogs_unit(i) for i in range(36)]
row("COGS producto (€)", cogs_prod)
stripe = [round(total_rev[i] * 0.014) for i in range(36)]
row("Comisión pasarela (€)", stripe)
pack = [all_units[i] * 40 for i in range(36)]
row("Packaging + CS variable (€)", pack)
total_cogs = [cogs_prod[i] + stripe[i] + pack[i] for i in range(36)]
row("TOTAL COGS", total_cogs, bold_row=True)

margin = [total_rev[i] - total_cogs[i] for i in range(36)]
row("MARGEN BRUTO", margin, bold_row=True, color=LIGHT_BG)

row("% Margen bruto", [round(margin[i]/total_rev[i]*100,1) if total_rev[i] else 0 for i in range(36)])

row("", ["" for _ in range(36)])

# OpEx
mkt_variable = [all_units[i] * cac(i) for i in range(36)]
row("Marketing variable (CAC × uds) (€)", mkt_variable)
fixed_list = [fixed(i) for i in range(36)]
row("Costes fijos (personal + tools + etc.) (€)", fixed_list)

# Inversión inicial en M1-M3
inv_m1 = 12000 + 8000 + 6500 + 1200 + 9000  # total 36700
# Spread over M1-M3
inv_schedule = [0]*36
inv_schedule[0] = round(inv_m1 * 0.5)
inv_schedule[1] = round(inv_m1 * 0.3)
inv_schedule[2] = inv_m1 - inv_schedule[0] - inv_schedule[1]
row("Inversión inicial (€)", inv_schedule)

total_opex = [mkt_variable[i] + fixed_list[i] + inv_schedule[i] for i in range(36)]
row("TOTAL OPEX", total_opex, bold_row=True)

row("", ["" for _ in range(36)])

ebitda = [margin[i] - total_opex[i] for i in range(36)]
row("EBITDA (€)", ebitda, bold_row=True, color="FFF9DD")
row("% EBITDA / Ingresos", [round(ebitda[i]/total_rev[i]*100,1) if total_rev[i] else 0 for i in range(36)])

# cumulative EBITDA
cum = []; acc = 0
for v in ebitda: acc += v; cum.append(acc)
row("EBITDA acumulado (€)", cum, bold_row=True, color=LIGHT_BG)

# Write to sheet
for ri, (label, values, b, color) in enumerate(rows_data, start=5):
    cc = pnl.cell(row=ri, column=1, value=label)
    if b: cc.font = bold
    if color: cc.fill = PatternFill("solid", fgColor=color)
    for i, v in enumerate(values, start=2):
        cell = pnl.cell(row=ri, column=i, value=v)
        if b and v != "": cell.font = bold
        if color: cell.fill = PatternFill("solid", fgColor=color)
        if isinstance(v, (int, float)) and label not in ("Juegos vendidos (uds)", "Precio medio venta (€)", "% Margen bruto", "% EBITDA / Ingresos"):
            cell.number_format = '#,##0'
        elif label == "Precio medio venta (€)" and isinstance(v, (int,float)):
            cell.number_format = '#,##0'
        elif "%" in label and isinstance(v, (int, float)):
            cell.number_format = '0.0"%"'
    # Yearly totals
    if label and label not in ("", "Precio medio venta (€)", "% Margen bruto", "% EBITDA / Ingresos"):
        for yi, (start_i, end_i, col) in enumerate([(0,12,38),(12,24,39),(24,36,40)]):
            vals = values[start_i:end_i]
            if all(isinstance(x,(int,float)) for x in vals):
                total = sum(vals)
                c = pnl.cell(row=ri, column=col, value=total)
                c.number_format = '#,##0'
                c.fill = PatternFill("solid", fgColor="FFF9DD")
                if b: c.font = bold

pnl.column_dimensions["A"].width = 42
for i in range(2, 41):
    pnl.column_dimensions[get_column_letter(i)].width = 11
pnl.freeze_panes = "B5"

# =====================================================================
# Sheet: Cash Flow
# =====================================================================
cf = wb.create_sheet("Cash Flow")
cf["A1"] = "FLUJO DE CAJA — MENSUAL 36 MESES"
cf["A1"].font = Font(bold=True, size=14, color=DARK)
cf["A2"] = "Caja inicial = aporte fundador 8.000 € + tramo ENISA 75.000 € en M1"
cf["A2"].font = italic

cf.cell(row=4, column=1, value="Concepto").font = header_font
cf.cell(row=4, column=1).fill = dark_fill
for i, m in enumerate(months, start=2):
    c = cf.cell(row=4, column=i, value=m)
    c.font = header_font; c.fill = dark_fill; c.alignment = center

def w(row_i, label, vals, b=False, color=None):
    cc = cf.cell(row=row_i, column=1, value=label)
    if b: cc.font = bold
    if color: cc.fill = PatternFill("solid", fgColor=color)
    for i, v in enumerate(vals, start=2):
        cell = cf.cell(row=row_i, column=i, value=v)
        if b and v != "": cell.font = bold
        if color: cell.fill = PatternFill("solid", fgColor=color)
        if isinstance(v, (int, float)): cell.number_format = '#,##0'

# Assume revenue collected same month, COGS paid with 30-day lag
cogs_paid = [0] + cogs_prod[:-1]  # shift
cash_in = [total_rev[i] for i in range(36)]
cash_out = [cogs_paid[i] + stripe[i] + pack[i] + mkt_variable[i] + fixed_list[i] + inv_schedule[i] for i in range(36)]
# ENISA disbursement
enisa = [0]*36; enisa[0] = 75000
# ENISA repayment (simplified: grace 12 m, then level payments)
enisa_pay = [0]*36
monthly_rate = 0.06/12
term_after_grace = 72
principal = 75000
pay = principal * monthly_rate / (1 - (1+monthly_rate)**-term_after_grace)
for i in range(12, 36):
    enisa_pay[i] = round(pay)

net = [cash_in[i] - cash_out[i] + enisa[i] - enisa_pay[i] for i in range(36)]

w(5, "Cash in (Ingresos cobrados)", cash_in)
w(6, "Cash out (COGS + OPEX + Inversión)", [-x for x in cash_out])
w(7, "Desembolso ENISA", enisa)
w(8, "Cuota ENISA", [-x for x in enisa_pay])
w(9, "FLUJO NETO MES", net, b=True, color="FFF9DD")

# Cumulative cash with initial 8000 founder
cash_balance = []; bal = 8000
for v in net: bal += v; cash_balance.append(bal)
w(10, "CAJA ACUMULADA", cash_balance, b=True, color=LIGHT_BG)

cf.column_dimensions["A"].width = 42
for i in range(2, 38):
    cf.column_dimensions[get_column_letter(i)].width = 11
cf.freeze_panes = "B5"

# =====================================================================
# Sheet: Resumen
# =====================================================================
sm = wb.create_sheet("Resumen", 0)
sm["A1"] = "FORGED WHEELS — RESUMEN FINANCIERO"
sm["A1"].font = Font(bold=True, size=18, color=DARK)
sm["A2"] = "SignalWaves SLU — CIF B19996313 — Barcelona"
sm["A2"].font = italic

sm["A4"] = "KPIs anuales"
sm["A4"].font = Font(bold=True, size=13, color=GOLD)

header_row = ["Métrica", "Año 1 (Jun26-May27)", "Año 2 (Jun27-May28)", "Año 3 (Jun28-May29)"]
for i,h in enumerate(header_row, start=1):
    c = sm.cell(row=5, column=i, value=h)
    c.font = header_font; c.fill = dark_fill; c.alignment = center; c.border = border

rows_summary = [
    ("Juegos vendidos", y1_units, y2_units, y3_units, "#,##0"),
    ("Precio medio venta (€)", 1800, 2000, 2200, '#,##0 "€"'),
    ("Ingresos B2C (€)", sum(revenue[:12]), sum(revenue[12:24]), sum(revenue[24:]), '#,##0 "€"'),
    ("Ingresos Lead-gen (€)", sum(leadgen_mo(i) for i in range(12)), sum(leadgen_mo(i) for i in range(12,24)), sum(leadgen_mo(i) for i in range(24,36)), '#,##0 "€"'),
    ("Ingresos SaaS (€)", 0, sum(saas_mo(i) for i in range(12,24)), sum(saas_mo(i) for i in range(24,36)), '#,##0 "€"'),
    ("INGRESOS TOTALES (€)", sum(total_rev[:12]), sum(total_rev[12:24]), sum(total_rev[24:]), '#,##0 "€"'),
    ("COGS totales (€)", sum(total_cogs[:12]), sum(total_cogs[12:24]), sum(total_cogs[24:]), '#,##0 "€"'),
    ("MARGEN BRUTO (€)", sum(margin[:12]), sum(margin[12:24]), sum(margin[24:]), '#,##0 "€"'),
    ("% Margen bruto", sum(margin[:12])/sum(total_rev[:12])*100, sum(margin[12:24])/sum(total_rev[12:24])*100, sum(margin[24:])/sum(total_rev[24:])*100, '0.0"%"'),
    ("Marketing (CAC × uds) (€)", sum(mkt_variable[:12]), sum(mkt_variable[12:24]), sum(mkt_variable[24:]), '#,##0 "€"'),
    ("Costes fijos (€)", sum(fixed_list[:12]), sum(fixed_list[12:24]), sum(fixed_list[24:]), '#,##0 "€"'),
    ("Inversión inicial (€)", sum(inv_schedule[:12]), 0, 0, '#,##0 "€"'),
    ("EBITDA (€)", sum(ebitda[:12]), sum(ebitda[12:24]), sum(ebitda[24:]), '#,##0 "€"'),
    ("% EBITDA", sum(ebitda[:12])/sum(total_rev[:12])*100 if sum(total_rev[:12]) else 0, sum(ebitda[12:24])/sum(total_rev[12:24])*100, sum(ebitda[24:])/sum(total_rev[24:])*100, '0.0"%"'),
    ("Caja final año (€)", cash_balance[11], cash_balance[23], cash_balance[35], '#,##0 "€"'),
]

for ri, (label, y1v, y2v, y3v, fmt) in enumerate(rows_summary, start=6):
    sm.cell(row=ri, column=1, value=label).font = bold if "TOTAL" in label or "EBITDA" in label.upper() or "MARGEN" in label.upper() else Font(size=11)
    for ci, v in enumerate([y1v, y2v, y3v], start=2):
        c = sm.cell(row=ri, column=ci, value=v)
        c.number_format = fmt
        c.alignment = right
        c.border = border
        if "TOTAL" in label or "EBITDA" in label or "MARGEN" in label:
            c.font = bold
            c.fill = light_fill

sm["A22"] = "Necesidad de financiación y uso de fondos"
sm["A22"].font = Font(bold=True, size=13, color=GOLD)

fund_rows = [
    ("Préstamo ENISA solicitado", 75000),
    ("Aporte propio fundador", 8000),
    ("TOTAL FUENTES", 83000),
    ("", ""),
    ("Homologación TÜV (3 SKUs)", 12000),
    ("Desarrollo producto (checkout, CRM)", 8000),
    ("Samples + tooling piloto", 6500),
    ("Marca OEPM + EUIPO", 1200),
    ("Marketing lanzamiento (M1-M3)", 9000),
    ("Fondo maniobra (12 meses operativos)", 46300),
    ("TOTAL USOS", 83000),
]
for ri, (label, val) in enumerate(fund_rows, start=23):
    sm.cell(row=ri, column=1, value=label)
    if val != "":
        c = sm.cell(row=ri, column=2, value=val)
        c.number_format = '#,##0 "€"'
        c.alignment = right
    if "TOTAL" in label:
        sm.cell(row=ri, column=1).font = bold
        sm.cell(row=ri, column=2).font = bold
        sm.cell(row=ri, column=1).fill = light_fill
        sm.cell(row=ri, column=2).fill = light_fill

sm.column_dimensions["A"].width = 50
for i in range(2, 5):
    sm.column_dimensions[get_column_letter(i)].width = 22

# =====================================================================
# Sheet: Inversión ENISA (detail)
# =====================================================================
inv = wb.create_sheet("Inversión ENISA")
inv["A1"] = "APLICACIÓN DE LOS FONDOS ENISA (75.000 €)"
inv["A1"].font = Font(bold=True, size=14, color=DARK)

inv_header = ["Partida", "Importe (€)", "% del préstamo", "Mes desembolso", "Justificación"]
for i,h in enumerate(inv_header, start=1):
    c = inv.cell(row=3, column=i, value=h)
    c.font = header_font; c.fill = dark_fill; c.alignment = center; c.border = border

inv_rows = [
    ("Homologación TÜV Teilegutachten (3 SKUs propios)", 12000, "16%", "M2-M4", "Requisito para venta road-legal catálogo propio en DE"),
    ("Desarrollo producto: checkout Stripe, CRM, cuenta cliente", 8000, "11%", "M1-M3", "Completar MVP para conversión comercial"),
    ("Samples de pre-producción (5 juegos sample)", 6500, "9%", "M1-M2", "Control de calidad antes de pedidos comerciales"),
    ("Registro Marca: OEPM (España) + EUIPO (UE)", 1200, "2%", "M1", "Protección de la marca 'Forged Wheels' en 27 países"),
    ("Marketing de lanzamiento (Meta + Google + PR) M1-M3", 9000, "12%", "M1-M3", "Campaña primeros 90 días para alcanzar break-even M9"),
    ("Salario fundador M4-M15 (12 meses a 2.500 €/mes)", 30000, "40%", "M4-M15", "Dedicación 100% al proyecto mientras no hay EBITDA positivo"),
    ("Herramientas + infraestructura 18 meses", 6300, "8%", "Mensual", "Vercel, fal.ai, Stripe, Plausible, gestoría, seguros RC"),
    ("Contingencia operativa", 2000, "3%", "Reserva", "Variaciones tipo cambio, imprevistos logística"),
    ("TOTAL", 75000, "100%", "", ""),
]
for ri, (label, imp, pct, mes, just) in enumerate(inv_rows, start=4):
    inv.cell(row=ri, column=1, value=label)
    c = inv.cell(row=ri, column=2, value=imp); c.number_format = '#,##0 "€"'; c.alignment = right
    inv.cell(row=ri, column=3, value=pct).alignment = right
    inv.cell(row=ri, column=4, value=mes)
    inv.cell(row=ri, column=5, value=just)
    if label == "TOTAL":
        for col in range(1,6):
            inv.cell(row=ri, column=col).font = bold
            inv.cell(row=ri, column=col).fill = light_fill

inv.column_dimensions["A"].width = 48
inv.column_dimensions["B"].width = 14
inv.column_dimensions["C"].width = 14
inv.column_dimensions["D"].width = 16
inv.column_dimensions["E"].width = 55

# Save
OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"Saved: {OUT}")
